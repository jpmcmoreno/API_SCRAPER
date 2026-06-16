# -*- coding: utf-8 -*-
"""
Web service del monitoreo de noticias.

GET  /noticias?analista=ANDREA,LUISA&desde=2026-06-01&hasta=2026-06-10&tipo=excel|json
     - Sin fechas: ultimos 2 dias. Sin tipo: excel (descarga directa).
     - formato=consolidado -> una hoja por fuente (formato de la Llamada a la API).
     - pais=COLOMBIA  -> filtra solo las fuentes de ese pais.
     - En el campo "analista" tambien se puede escribir un PAIS (ej: COLOMBIA):
       se descargan las noticias de las fuentes de ese pais.
GET  /            formulario web
GET  /salud       estado y total de noticias
POST /cargar      carga de noticias (JSON, requiere header X-API-Key)
"""
import io
import os
import re
import json
import sqlite3
import unicodedata
from datetime import datetime, timedelta

from flask import Flask, request, jsonify, send_file, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

app = Flask(__name__)
app.url_map.strict_slashes = False   # /noticias y /noticias/ funcionan igual


@app.after_request
def _cors(resp):
    # permite que la interfaz en GitHub Pages consuma esta API desde el navegador
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type, X-API-Key"
    resp.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return resp


@app.before_request
def _ruta_minusculas():
    # tolera /Noticias, /SALUD, etc.
    from flask import request as _rq
    if _rq.path != _rq.path.lower() and _rq.path.lower() in ("/noticias", "/salud", "/cargar", "/"):
        from werkzeug.routing import RequestRedirect
        raise RequestRedirect(_rq.path.lower() + (("?" + _rq.query_string.decode()) if _rq.query_string else ""))

RUTA_BD = os.environ.get("RUTA_BD", os.path.join(os.path.dirname(__file__), "monitoreo.db"))
API_KEY = os.environ.get("API_KEY", "")   # definir en Render

# ----------------------------- MAPA FUENTE -> PAIS --------------------------
# fuentes_pais.json: { "<fuente>": "<PAIS>", ... }  (generado desde el Excel).
RUTA_PAIS = os.path.join(os.path.dirname(__file__), "fuentes_pais.json")
try:
    with open(RUTA_PAIS, encoding="utf-8") as _fh:
        FUENTE_PAIS = json.load(_fh)
except Exception:
    FUENTE_PAIS = {}


def _norm(s):
    """Normaliza para comparar paises: sin acentos, sin signos, MAYUS."""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Za-z0-9]", "", s).upper()


# indices auxiliares: PAIS_NORM -> {fuentes}, PAIS_NORM -> nombre bonito
PAIS_FUENTES, PAIS_CANON = {}, {}
for _f, _p in FUENTE_PAIS.items():
    _k = _norm(_p)
    if not _k:
        continue
    PAIS_FUENTES.setdefault(_k, set()).add(_f)
    PAIS_CANON.setdefault(_k, str(_p))
PAISES_NORM = set(PAIS_FUENTES.keys())


def _fuentes_de_paises(paises_norm):
    fu = set()
    for p in paises_norm:
        fu |= PAIS_FUENTES.get(p, set())
    return sorted(fu)


ESQUEMA = """
CREATE TABLE IF NOT EXISTS noticias (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    analista TEXT NOT NULL,
    fuente   TEXT NOT NULL,
    titular  TEXT,
    link     TEXT NOT NULL UNIQUE,
    fecha    TEXT,
    cuerpo   TEXT,
    cargado_en TEXT DEFAULT (datetime('now'))
);
CREATE INDEX IF NOT EXISTS idx_fecha ON noticias(fecha);
CREATE INDEX IF NOT EXISTS idx_analista ON noticias(analista);
CREATE INDEX IF NOT EXISTS idx_fuente ON noticias(fuente);
"""


def conectar():
    con = sqlite3.connect(RUTA_BD, timeout=30)
    con.executescript(ESQUEMA)
    return con


# ----------------------------- consultas ------------------------------------
def filtrar(analista, desde, hasta, pais=""):
    """
    Construye el SQL del filtro.
    - El campo 'analista' acepta nombres de analista Y nombres de pais mezclados
      (separados por coma). Los paises se traducen a sus fuentes.
    - 'pais' (parametro aparte / desplegable) restringe ademas por pais (AND).
    """
    cond, params = [], []

    # 1) separar el campo "analista" en analistas reales y paises escritos ahi
    analistas, paises_en_campo = [], []
    for a in (analista or "").split(","):
        a = a.strip()
        if not a:
            continue
        if _norm(a) in PAISES_NORM:
            paises_en_campo.append(_norm(a))
        else:
            analistas.append(a.upper())

    # grupo "quien": analistas OR fuentes-del-pais-escrito  (union)
    piezas = []
    if analistas:
        piezas.append("UPPER(analista) IN (%s)" % ",".join("?" * len(analistas)))
        params += analistas
    if paises_en_campo:
        fu = _fuentes_de_paises(paises_en_campo)
        if fu:
            piezas.append("UPPER(fuente) IN (%s)" % ",".join("?" * len(fu)))
            params += [f.upper() for f in fu]
        else:
            piezas.append("1=0")
    if piezas:
        cond.append("(" + " OR ".join(piezas) + ")")

    # 2) filtro de pais explicito (desplegable) -> AND
    paises_param = [_norm(p) for p in (pais or "").split(",")
                    if p.strip() and _norm(p) in PAISES_NORM]
    if paises_param:
        fu = _fuentes_de_paises(paises_param)
        if fu:
            cond.append("UPPER(fuente) IN (%s)" % ",".join("?" * len(fu)))
            params += [f.upper() for f in fu]
        else:
            cond.append("1=0")

    # 3) fechas
    cond.append("fecha >= ?"); params.append(desde + " 00:00:00")
    cond.append("fecha <= ?"); params.append(hasta + " 23:59:59")

    sql = ("SELECT analista, fuente, titular, link, fecha, cuerpo FROM noticias "
           "WHERE " + " AND ".join(cond) + " ORDER BY analista, fecha DESC")
    return sql, params


def encabezado(ws, columnas, anchos):
    ws.append(columnas)
    azul = PatternFill("solid", start_color="1F3864")
    for c in range(1, len(columnas) + 1):
        cel = ws.cell(row=1, column=c)
        cel.font = Font(bold=True, color="FFFFFF", name="Arial")
        cel.fill = azul
    for col, ancho in zip("ABCDEFGH", anchos):
        ws.column_dimensions[col].width = ancho
    ws.freeze_panes = "A2"


def excel_plano(filas):
    wb = Workbook(); ws = wb.active; ws.title = "Noticias"
    encabezado(ws, ["ANALISTA", "FUENTE", "PAIS", "TITULAR", "LINK", "FECHA", "CUERPO DE LA NOTICIA"],
               (12, 18, 14, 60, 50, 20, 120))
    for analista, fuente, titular, link, fecha, cuerpo in filas:
        ws.append([analista, fuente, FUENTE_PAIS.get(fuente, ""), titular, link, fecha, cuerpo])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


def excel_consolidado(filas):
    # Formato de la Llamada a la API: una hoja por fuente (no se toca).
    wb = Workbook(); wb.remove(wb.active)
    for analista, fuente, titular, link, fecha, cuerpo in filas:
        hoja = str(fuente or "sin_fuente")[:31]
        if hoja not in [w.title for w in wb.worksheets]:
            ws = wb.create_sheet(hoja)
            encabezado(ws, ["TITULAR", "LINK", "FECHA", "CUERPO DE LA NOTICIA"],
                       (60, 50, 22, 120))
        wb[hoja].append([titular, link, fecha, cuerpo])
    if not wb.worksheets:
        ws = wb.create_sheet("Noticias")
        encabezado(ws, ["TITULAR", "LINK", "FECHA", "CUERPO DE LA NOTICIA"], (60, 50, 22, 120))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


@app.get("/noticias")
def noticias():
    hoy = datetime.utcnow().date()
    desde = request.args.get("desde") or str(hoy - timedelta(days=2))
    hasta = request.args.get("hasta") or str(hoy)
    tipo = (request.args.get("tipo") or "excel").lower()
    formato = (request.args.get("formato") or "consolidado").lower()
    pais = request.args.get("pais", "")

    sql, params = filtrar(request.args.get("analista", ""), desde, hasta, pais)
    con = conectar()
    filas = con.execute(sql, params).fetchall()
    con.close()

    if tipo == "json":
        claves = ["analista", "fuente", "titular", "link", "fecha", "cuerpo"]
        noticias = []
        for f in filas:
            d = dict(zip(claves, f))
            d["pais"] = FUENTE_PAIS.get(f[1], "")
            noticias.append(d)
        return jsonify({"total": len(filas), "desde": desde, "hasta": hasta,
                        "pais": pais, "noticias": noticias})

    buf = excel_plano(filas) if formato == "plano" else excel_consolidado(filas)
    nombre = f"Noticias_{desde}_a_{hasta}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nombre,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ----------------------------- carga remota ---------------------------------
@app.post("/cargar")
def cargar():
    if not API_KEY or request.headers.get("X-API-Key") != API_KEY:
        return jsonify({"error": "API key invalida"}), 401
    datos = request.get_json(silent=True)
    if not isinstance(datos, list):
        return jsonify({"error": "se espera una lista JSON de noticias"}), 400

    con = conectar(); cur = con.cursor()
    nuevas = mejoradas = 0
    for n in datos:
        link = str(n.get("link", "")).strip()
        if not link.lower().startswith("http"):
            continue
        cuerpo = str(n.get("cuerpo") or "")
        existe = cur.execute("SELECT LENGTH(cuerpo) FROM noticias WHERE link=?", (link,)).fetchone()
        if existe is None:
            cur.execute("INSERT INTO noticias (analista, fuente, titular, link, fecha, cuerpo) "
                        "VALUES (?,?,?,?,?,?)",
                        (str(n.get("analista") or ""), str(n.get("fuente") or ""),
                         str(n.get("titular") or ""), link, n.get("fecha"), cuerpo))
            nuevas += 1
        elif len(cuerpo) > (existe[0] or 0):
            cur.execute("UPDATE noticias SET cuerpo=?, fecha=COALESCE(fecha,?) WHERE link=?",
                        (cuerpo, n.get("fecha"), link))
            mejoradas += 1
    con.commit()
    total = cur.execute("SELECT COUNT(*) FROM noticias").fetchone()[0]
    con.close()
    return jsonify({"recibidas": len(datos), "nuevas": nuevas,
                    "mejoradas": mejoradas, "total_en_base": total})


# ----------------------------- estado y portada -----------------------------
@app.get("/salud")
def salud():
    con = conectar()
    total, ultima = con.execute("SELECT COUNT(*), MAX(cargado_en) FROM noticias").fetchone()
    con.close()
    return jsonify({"ok": True, "total_noticias": total, "ultima_carga": ultima,
                    "paises_disponibles": sorted(PAIS_CANON.values()),
                    "fuentes_con_pais": len(FUENTE_PAIS)})


@app.get("/")
def portada():
    con = conectar()
    total, = con.execute("SELECT COUNT(*) FROM noticias").fetchone()
    con.close()
    opciones_pais = "".join(
        f'<option value="{p}">{p}</option>' for p in sorted(PAIS_CANON.values()))
    return Response(f"""<!doctype html><html lang="es"><head><meta charset="utf-8">
<title>Monitoreo de noticias</title>
<style>body{{font-family:Segoe UI,Arial;background:#f4f6fb;padding:40px}}
.caja{{max-width:540px;margin:auto;background:#fff;border-radius:12px;padding:28px 32px;
box-shadow:0 2px 10px rgba(31,56,100,.15)}}h1{{color:#1F3864;font-size:21px;margin-top:0}}
label{{display:block;margin:10px 0 4px;font-weight:600;color:#2E4057}}
input,select{{width:100%;padding:8px;border:1px solid #c5cede;border-radius:6px;box-sizing:border-box}}
button{{margin-top:16px;background:#1F3864;color:#fff;border:0;border-radius:6px;
padding:10px 22px;cursor:pointer}}.nota{{color:#667;font-size:12px;margin-top:6px}}</style>
</head><body><div class="caja"><h1>Monitoreo de noticias</h1>
<form action="/noticias" method="get">
<label>Analistas o país (vacío = todos)</label>
<input name="analista" placeholder="ANDREA,LUISA  ·  o un país: COLOMBIA">
<p class="nota">Puedes escribir nombres de analista, un país (ej: COLOMBIA), o ambos separados por coma.</p>
<label>País (opcional, restringe por país)</label>
<select name="pais"><option value="">— Todos los países —</option>{opciones_pais}</select>
<label>Desde (vacío = hace 2 días)</label><input type="date" name="desde">
<label>Hasta (vacío = hoy)</label><input type="date" name="hasta">
<label>Tipo</label><select name="tipo"><option value="excel">Excel</option><option value="json">JSON</option></select>
<label>Formato del Excel</label><select name="formato">
<option value="consolidado">Consolidado para la API (hoja por fuente)</option>
<option value="plano">Plano (una sola hoja, incluye país)</option></select>
<button type="submit">Descargar</button></form>
<p class="nota">{total:,} noticias en la base · {len(FUENTE_PAIS)} fuentes con país.</p></div></body></html>""",
        mimetype="text/html")


@app.errorhandler(404)
def _no_encontrado(e):
    return jsonify({"error": "ruta no encontrada",
                    "rutas_validas": ["/", "/noticias", "/salud", "/cargar (POST)"],
                    "ejemplo": "/noticias?analista=JUAN&desde=2026-06-01&hasta=2026-06-10",
                    "ejemplo_pais": "/noticias?pais=COLOMBIA   o   /noticias?analista=COLOMBIA"}), 404


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8000)))
